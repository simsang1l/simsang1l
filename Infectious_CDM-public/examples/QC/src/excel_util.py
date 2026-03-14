import pandas as pd
from openpyxl import load_workbook

def write_df_to_excel(excel_path, sheetname, df):
    "원하는 엑셀 파일에 데이터 쓰기"
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode = "a", if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row , index=False, header=None)

def get_sheet_row_count(excel_path, sheetname):
    wb = load_workbook(excel_path)
    ws = wb[sheetname]

    row_count = ws.max_row

    return row_count

def find_excel_row_and_write_error_ratio(excel_path, sheetname, key_value, error_count, row_count, error_rate):
    """
    key_column이 A1에 있다는 가정으로 실행
    """
    wb = load_workbook(excel_path)
    ws = wb[sheetname]

    for row in ws.iter_rows(min_row=2, max_col=ws.max_column): # 첫 행은 헤더로 가정, 데이터는 두 번째 행부터 시작
        if row[0].value == key_value:
            error_count_cell_address = 'G' + str(row[0].row)  # G열 같은 행의 셀 주소 생성
            row_count_cell_address = 'H' + str(row[0].row)  # G열 같은 행의 셀 주소 생성
            error_rate_cell_address = 'I' + str(row[0].row)  # G열 같은 행의 셀 주소 생성

            ws[error_count_cell_address] = error_count  # 새로운 값 할당
            ws[row_count_cell_address] = row_count  # 새로운 값 할당
            ws[error_rate_cell_address] = error_rate  # 새로운 값 할당
            break
    
    wb.save(excel_path)

def find_excel_row_and_write_metadata_count(excel_path, sheetname, key_value, feature_count, row_count, feature_patient_count, patient_count, patient_ratio):
    """
    key_column이 A1에 있다는 가정으로 실행
    """
    wb = load_workbook(excel_path)
    ws = wb[sheetname]

    for row in ws.iter_rows(min_row=2, max_col=ws.max_column): # 첫 행은 헤더로 가정, 데이터는 두 번째 행부터 시작
        if row[0].value == key_value:
            feature_count_cell_address = 'G' + str(row[0].row)  # G열 같은 행의 셀 주소 생성
            row_count_cell_address = 'H' + str(row[0].row)  # H열 같은 행의 셀 주소 생성
            feature_patient_count_cell_address = 'I' + str(row[0].row)  # I열 같은 행의 셀 주소 생성
            patient_count_cell_address = 'J' + str(row[0].row)  # J열 같은 행의 셀 주소 생성
            patient_ratio_cell_address = 'K' + str(row[0].row)  # J열 같은 행의 셀 주소 생성

            ws[feature_count_cell_address] = feature_count  # 새로운 값 할당
            ws[row_count_cell_address] = row_count  # 새로운 값 할당
            ws[feature_patient_count_cell_address] = feature_patient_count  # 새로운 값 할당
            ws[patient_count_cell_address] = patient_count  # 새로운 값 할당
            ws[patient_ratio_cell_address] = patient_ratio  # 새로운 값 할당
            break
    
    wb.save(excel_path)
                    
